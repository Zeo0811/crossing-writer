from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, Optional
from datetime import datetime, date
from openpyxl import load_workbook
from .sanitize import sanitize_filename

@dataclass
class XlsxRow:
    account: str
    title: str
    published_at: str
    is_original: bool
    position: Optional[int]
    url: str
    cover: Optional[str]
    author: Optional[str]
    summary: Optional[str]
    xlsx_row_number: int

def _to_date_str(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    # fallback parse YYYY-MM-DD prefix
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None

def iter_xlsx_rows(xlsx_path: Path) -> Iterator[XlsxRow]:
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return
    for idx, row in enumerate(rows, start=2):
        row = list(row) + [None] * (9 - len(row))
        account, title, pub, pos, orig, url, cover, author, summary = row[:9]
        if not title or not url:
            continue
        date_str = _to_date_str(pub)
        if not date_str:
            continue
        yield XlsxRow(
            account=str(account).strip() if account else "",
            title=str(title).strip(),
            published_at=date_str,
            is_original=(str(orig).strip() == "是") if orig else False,
            position=int(pos) if pos is not None and str(pos).strip().isdigit() else None,
            url=str(url).strip(),
            cover=str(cover).strip() if cover else None,
            author=str(author).strip().replace("&nbsp;", " ") if author else None,
            summary=str(summary).strip() if summary else None,
            xlsx_row_number=idx,
        )

def _levenshtein(a: str, b: str) -> int:
    if a == b: return 0
    if not a: return len(b)
    if not b: return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(min(
                curr[-1] + 1,
                prev[j] + 1,
                prev[j-1] + (0 if ca == cb else 1)
            ))
        prev = curr
    return prev[-1]

def find_html(html_dir: Path, row: XlsxRow) -> Optional[Path]:
    expected_stem = f"{row.published_at}_{sanitize_filename(row.title)}"
    exact = html_dir / f"{expected_stem}.html"
    if exact.exists():
        return exact
    # fuzzy: scan same-date files
    prefix = f"{row.published_at}_"
    candidates = [p for p in html_dir.glob(f"{prefix}*.html")]
    if not candidates:
        return None
    san_title = sanitize_filename(row.title)
    best = None
    best_d = 10**9
    for c in candidates:
        stem_title = c.stem[len(prefix):]
        d = _levenshtein(san_title, stem_title)
        if d < best_d:
            best_d = d; best = c
    if best is not None and best_d <= 5:
        return best
    return None
